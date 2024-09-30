from openpyxl import load_workbook
from openpyxl import Workbook
import os

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

nome_arquivo = "C:\\Users\\luanm\\OneDrive\\Desktop\\Sites\\Estudos\\estudos python\\Python RPA\\Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

#Seleciona a sheet de Dados
sheet_selecionada = planilha_aberta['Dados']

criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

#Criando preenchimento cor Azul
corAzul = PatternFill(start_color='0099CCFF',
                    end_color='0099CCFF',
                    fill_type='solid')

#Criando preenchimento cor Amarelo
corAmarelo = PatternFill(start_color='00FFFFCC',
                    end_color='00FFFFCC',
                    fill_type='solid')

#Coloca bordar preta na celula
bdFina = Side(border_style="thin", color="000000")
borda = Border(left=bdFina, right=bdFina, top=bdFina, bottom=bdFina)

for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
                    
        #Preenche os dados
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
        #Colocando cores nas celulas
        selecionaSheetVendasNovaPlanilha[celulaColunaA].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[celulaColunaB].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[celulaColunaC].fill = corAmarelo
        
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
        
    else:
        
        #Adiciona o nome do funcionario que esta na linha que o código está passando
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        nova_planilha = criandoNovoArquivoExcel.active
        
        nova_planilha.title = "Vendas"
        
        caminhoNovaPlanilha = "C:\\Users\\luanm\\OneDrive\\Desktop\\Sites\\Estudos\\estudos python\\Python RPA\\" + sheet_selecionada['A%s' % linha].value + ".xlsx"
        
        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']
        
        #Coloca os titulos
        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"
        
        #Preenche as informações na segunda linha
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value
        
        #Colocando cores nas celulas
        selecionaSheetVendasNovaPlanilha['A1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['B1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['C1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['A1'].border = borda
        selecionaSheetVendasNovaPlanilha['B1'].border = borda
        selecionaSheetVendasNovaPlanilha['C1'].border = borda
        
        selecionaSheetVendasNovaPlanilha['A2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['B2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['C2'].fill = corAmarelo
        
        #Aumentando o tamanho das colunas
        selecionaSheetVendasNovaPlanilha.column_dimensions["A"].width = 18
        selecionaSheetVendasNovaPlanilha.column_dimensions["B"].width = 25
        selecionaSheetVendasNovaPlanilha.column_dimensions["C"].width = 15
        
        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)
        
        
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)