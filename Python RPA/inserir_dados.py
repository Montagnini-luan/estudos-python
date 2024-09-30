from openpyxl import load_workbook
import os
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.cell import cell

caminho_nome_arquivo = 'C:\\Users\luanm\OneDrive\Desktop\Sites\Estudos\estudos python\Python RPA\InserirDados.xlsx'
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

sheet_selecionada = planilha_aberta['Aluno']


dadosTabela = [
    ['Nome', 'Idade'],
    ['Berenice', 28],
    ['Caio', 32],
    ['Nicole', 34],
    ['Leonardo', 19],
    ['Amanda', 25]
]

for linhaPlanilha in dadosTabela:
    sheet_selecionada.append(linhaPlanilha)

corTitulo = PatternFill(start_color='00FFFF00',
                        end_color='00FFFF00',
                        fill_type='solid')

corCelulas = PatternFill(start_color='00CCFFCC',
                        end_color='00CCFFCC',
                        fill_type='solid')


sheet_selecionada["A1"].fill = corTitulo
sheet_selecionada["B1"].fill = corTitulo

for linha in range(2, len(sheet_selecionada['A']) +1):

    celulaColunaA = "A" + str(linha)
    celulaColunaB = "B" + str(linha)

    sheet_selecionada[celulaColunaA].fill = corCelulas
    sheet_selecionada[celulaColunaB].fill = corCelulas

planilha_aberta.save(filename=caminho_nome_arquivo)

os.startfile(caminho_nome_arquivo)