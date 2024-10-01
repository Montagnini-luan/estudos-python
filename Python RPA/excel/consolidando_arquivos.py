import pandas as opcoesPandas
import os

caminho_arquivos = r"C:\Users\luanm\OneDrive\Desktop\Sites\Estudos\estudos python\Python RPA\excel"

lista_arquvio = os.listdir(caminho_arquivos)

lista_caminho_arquivo = [caminho_arquivos + '\\' + arquivo for arquivo in lista_arquvio if arquivo[-4:] == 'xlsx' ]

dados_arquivo = opcoesPandas.DataFrame()

for arquivo in lista_caminho_arquivo:

    dados = opcoesPandas.read_excel(arquivo)
    dados_arquivo = dados_arquivo._append(dados)

dados_arquivo.to_excel(r"C:\Users\luanm\OneDrive\Desktop\Sites\Estudos\estudos python\Python RPA\arquivo consolidado.xlsx")

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

caminho_arquivo_dados_sistema = "C:\\Users\\luanm\\OneDrive\\Desktop\\Sites\\Estudos\\estudos python\\Python RPA\\arquivo consolidado.xlsx"
planilha_dados_sistema = load_workbook(filename=caminho_arquivo_dados_sistema)

sheet_planilha_sistema = planilha_dados_sistema['Sheet1']

sheet_planilha_sistema.delete_cols(1)

sheet_planilha_sistema.title = 'Dados Consolidados'

sheet_planilha_sistema.column_dimensions['A'].width = 35
sheet_planilha_sistema.column_dimensions['B'].width = 40

cor_cinza = PatternFill(start_color='00FFFFFF',
                       end_color='00FFFFFF',
                       fill_type='solid')

cor_amarela = PatternFill(start_color='00FFFFCC',
                       end_color='00FFFFCC',
                       fill_type='solid')

bd_fina = Side(border_style='thin', color='000000')
borda = Border(left=bd_fina, right=bd_fina, top=bd_fina, bottom=bd_fina)


sheet_planilha_sistema['A1'].fill = cor_amarela
sheet_planilha_sistema['B1'].fill = cor_amarela
sheet_planilha_sistema['C1'].fill = cor_amarela

sheet_planilha_sistema['A1'].border = borda
sheet_planilha_sistema['B1'].border = borda
sheet_planilha_sistema['C1'].border = borda

for linha in range(2, len(sheet_planilha_sistema['A']) + 1):

    celula_coluna_a = 'A' + str(linha)
    celula_coluna_b = 'B' + str(linha)
    celula_coluna_c = 'C' + str(linha)

    sheet_planilha_sistema[celula_coluna_a].fill = cor_cinza
    sheet_planilha_sistema[celula_coluna_b].fill = cor_cinza
    sheet_planilha_sistema[celula_coluna_c].fill = cor_cinza

    sheet_planilha_sistema[celula_coluna_a].border = borda
    sheet_planilha_sistema[celula_coluna_b].border = borda
    sheet_planilha_sistema[celula_coluna_c].border = borda


ultima_linha = linha + 1
print(ultima_linha)

celula_ultima_linha = 'C' + str(ultima_linha)

formula_soma = "=SUM(C2:C" + str(linha) + ")"

sheet_planilha_sistema[celula_ultima_linha] = formula_soma

planilha_dados_sistema.save(filename=caminho_arquivo_dados_sistema)

os.startfile(caminho_arquivo_dados_sistema)

