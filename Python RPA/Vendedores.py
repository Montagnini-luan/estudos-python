from openpyxl import load_workbook
import os

caminho_nome_arquivo = 'C:\\Users\luanm\OneDrive\Desktop\Sites\Estudos\estudos python\Python RPA\Vendedores.xlsx'
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

sheet_selecionada = planilha_aberta['Vendas']

somarAmandaMartins = 0
somarElianeMoreira = 0
somarLeonardoAlmeida = 0
somarNicolasPereira = 0

for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    if sheet_selecionada['A%s' % linha].value == "Amanda Martins":
        somarAmandaMartins = somarAmandaMartins + sheet_selecionada['C%s' % linha].value
        
    elif sheet_selecionada['A%s' % linha].value == "Eliane Moreira":
        somarElianeMoreira = somarElianeMoreira + sheet_selecionada['C%s' % linha].value
        
    elif sheet_selecionada['A%s' % linha].value == "Nicolas Pereira":
        somarNicolasPereira = somarNicolasPereira + sheet_selecionada['C%s' % linha].value
        
    elif sheet_selecionada['A%s' % linha].value == "Leonardo Almeida":
        somarLeonardoAlmeida = somarLeonardoAlmeida + sheet_selecionada['C%s' % linha].value
    
sheet_resumo = planilha_aberta.create_sheet(title='Resumo')

sheet_resumo['A1'] = "Vendedores"
sheet_resumo['B1'] = "VENDAS"

sheet_resumo['A2'] = "AmandaMartins"
sheet_resumo['B2'] = somarAmandaMartins

sheet_resumo['A3'] = "ElianeMoreira"
sheet_resumo['B3'] = somarElianeMoreira

sheet_resumo['A4'] = "LeonardoAlmeida"
sheet_resumo['B4'] = somarLeonardoAlmeida

sheet_resumo['A5'] = "NicolasPereira"
sheet_resumo['B5'] = somarNicolasPereira

planilha_aberta.save(filename=caminho_nome_arquivo)

os.startfile(caminho_nome_arquivo)